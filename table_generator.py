from time import localtime, strftime
from openpyxl import Workbook
from os import listdir
from os.path import isfile, join
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
import argparse
from scipy.signal import argrelextrema
from math import floor

parser = argparse.ArgumentParser(description='Select directory')
parser.add_argument('comet_type', type=str, help='Enter jupiter or long or halley')
args = parser.parse_args()

wb = Workbook()
ws = wb.active

ws['A1'] = 'Comet'
ws['B1'] = 'Name'
ws['C1'] = 'Date'
ws['D1'] = 'Distance (AU)'
ws['E1'] = 'Period (yr)'
ws['F1'] = 'Semi-major axis (AU)'
ws['G1'] = 'Eccentricity'
ws['H1'] = 'Inclincation (DEG)'

mypath = '/Users/angelviolinist/NASA/emails/' + args.comet_type + '/'
onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
comet = []
name = []
date = []
distance = []
period = []
semi = []
ecc = []
incl = []
approaches_dict = {}

for f in onlyfiles:
    
    got_per = False
    got_ecc = False
    got_incl = False
    comet.append(f.replace('.txt',''))
    approaches_dict[f.replace('.txt','')] = []
    source = open(mypath + f,'r')
    for idx,line in enumerate(source):

        if 'EC= ' in line and not got_ecc:
            ecc.append(float(line.split()[1]))
            got_ecc = True
        elif 'IN= ' in line and not got_incl:
            incl.append(float(line.split()[5]))
            got_incl = True
        elif 'A= ' in line:
            semi.append(float(line.split()[1]))
        elif 'PER= ' in line and not got_per:
            try:
                period.append(float(line.split()[1]))
                got_per = True
            except:
                period.append(line.split()[1])
                got_per = True
        elif 'Target body name:' in line:
            find_name = line.split()[3]
            name.append(find_name)
            
        elif '$$SOE' in line:
            start = idx
        elif '$$EOE' in line:
            end = idx
            break
    
    source.seek(0)
    dates = []
    distances = []
    for idx,line in enumerate(source):
        if idx > start and idx < end:
            linesplit = line.split(', ')
            dates.append(linesplit[1].replace('A.D. ','').replace(' 00:00:00.0000',''))
            distances.append(float(linesplit[3].strip()))
    source.close()
    distances_array = np.array(distances)
    clos = argrelextrema(distances_array, np.less)
    close = ['']
    close[0] = []
    for i in clos[0]:
        if distances_array[i] < 1.5:
            close[0].append(i)
    for i in range(len(close[0])):
        approaches_dict[f.replace('.txt','')].append(dates[close[0][i]])
        approaches_dict[f.replace('.txt','')].append(distances_array[close[0][i]])

for i in range(len(onlyfiles)):
    current = ws.max_row + 1
    
    ws['A' + str(current)] = comet[i]
    ws['B' + str(current)] = name[i]
    ws['C' + str(current)] = approaches_dict[comet[i]][0]
    ws['D' + str(current)] = approaches_dict[comet[i]][1]
    ws['E' + str(current)] = period[i]
    ws['F' + str(current)] = semi[i]
    ws['G' + str(current)] = ecc[i]
    ws['H' + str(current)] = incl[i]
    
    if len(approaches_dict[comet[i]]) > 2:
        lock = ws.max_row
        for count,entry in enumerate(approaches_dict[comet[i]]):
            if count < 2:
                continue
            else:
                current = lock + floor(count / 2)
                ws['A' + str(current)] = comet[i]
                ws['B' + str(current)] = name[i]
                ws['E' + str(current)] = period[i]
                ws['F' + str(current)] = semi[i]
                ws['G' + str(current)] = ecc[i]
                ws['H' + str(current)] = incl[i]
                if count % 2 == 0:
                    ws['C' + str(current)] = approaches_dict[comet[i]][count]
                else:
                    ws['D' + str(current)] = approaches_dict[comet[i]][count]

dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))  
for col, value in dims.items():
    ws.column_dimensions[col].width = value
    
wb.save(args.comet_type + " comet table " + strftime("%Y-%m-%d %H:%M:%S", localtime()) + ".xlsx")