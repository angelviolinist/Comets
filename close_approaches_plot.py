import pandas as pd
import matplotlib.pyplot as plt
from os import listdir
from os.path import isfile, join
from math import modf

def format_time(raw):
    if len(str(raw)) < 2:
        out = '0' + str(raw)
    else:
        out = str(raw)
    return out

mypath = '/Users/angelviolinist/NASA/emails/'
onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]

# dictionary: key maps to name, time of close approach, close approach distance(AU)

database = {
    # format is like
    # '289P': ['Blanpain','2035 Nov 06.67816', .081704]
}

for f in onlyfiles:
    source = open(mypath + f,'r')
    targets = []
    for idx,line in enumerate(source):
        # get name from txt file
        if 'JPL/HORIZONS' in line:
            if '/' in line.split()[1]:
                linesplit = line.split()[1].split('/')
                name = linesplit[1]
                # make sure name doesn't have a number at the end
                check_name = line.split()[2]
                try:
                    pd.to_datetime(check_name)
                except:
                    name = name + ' ' + check_name
            else:
                name = line.split()[1]
        if 'Rec #:' in line:
            rec = line.split()[1].replace('#:','')
        if 'Time (JDTDB)' in line:
            row = idx
            break
    source.seek(0)
    for idx,line in enumerate(source):
        if idx == row + 2:
            linesplit = line.split()
            # get close approach distance
            dist = float(linesplit[6])
            
            # get and format the time of close approach correctly
            raw_date = linesplit[2] + '-' + linesplit[3] + '-' + linesplit[4]
            time = modf(float(raw_date[11:]) * 24)
            hour = int(time[1])
            minutes = modf(time[0] * 60)
            minute = int(minutes[1])
            second = round(minutes[0] * 60)
            string_date = raw_date[:11] + ' ' + format_time(hour) + ':' + format_time(minute) + ':' + format_time(second)
            date = [pd.to_datetime(string_date)]
            database[f.rstrip('.txt')] = [name, date, dist, string_date, rec]
            break

# plt.figure(figsize=(5,4))
label = []
for idx,key in enumerate(database.keys()):
    plt.plot_date(database[key][1], database[key][2])
    label.append(key + ' ' + database[key][0])

plt.title('All comets')
plt.xlabel('Dates')
plt.ylabel('Distance(AU)')
plt.yscale('log')
plt.legend(label,bbox_to_anchor=(1.0,1,0.005,0.005),loc='upper left')

plt.show()

# for key in database.keys():
#     print(key)
#     print(database[key][4])
#     print(database[key][3])

import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'Comet'
ws['B1'] = 'Name'
ws['C1'] = 'Date'
ws['D1'] = 'Distance(AU)'
for idx,key in enumerate(database.keys()):
    ws['A' + str(idx + 2)] = key
    ws['B' + str(idx + 2)] = database[key][0]
    ws['C' + str(idx + 2)] = database[key][3]
    ws['D' + str(idx + 2)] = database[key][2]
wb.save('table_approaches.xlsx')
wb.close()